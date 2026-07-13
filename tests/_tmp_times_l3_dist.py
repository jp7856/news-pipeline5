"""TIMES L1/L2/L3 현행(2024~) 분포 + basic.xlsx L3 분포."""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\jp\work\news-pipeline5")
import openpyxl
from agents.sub_agents.cefr_checker import LEVELS
from agents.sub_agents.article_classifier import classify

ARTICLES = r"C:\Users\jp\Desktop\기사\articles.xlsx"
BASIC    = r"C:\Users\jp\Desktop\기사\basic.xlsx"

EXCL = {
    "Photo News","Briefs","Star Brief","News in Brief","Cartoon",
    "Did You Know","Did You Know?","Debating","Cover","Q & A","NE You",
    "My Journal","Book Review","Stories","Story",
    "Readings for Junior","VoA Broadcast News","Think About It","My Diary",
    "Pros & Cons",
}
_SE      = re.compile(r"(?<=[.!?])\s+")
_YEAR_RE = re.compile(r"\b(20\d{2})\b")

def avg_sl(t):
    p = [len(re.findall(r"[A-Za-z']+", s)) for s in _SE.split(t.strip())]
    p = [w for w in p if w >= 1]
    return sum(p) / len(p) if p else 0.0

def pct(xs, p):
    s = sorted(xs)
    if not s: return 0.0
    idx = (len(s)-1)*p/100
    lo = int(idx); hi = min(lo+1, len(s)-1)
    return s[lo] + (s[hi]-s[lo])*(idx-lo)

# ── articles.xlsx 2024~ TIMES L1/L2/L3 ──────────────────────────────────
wb   = openpyxl.load_workbook(ARTICLES, read_only=True, data_only=True)
ws   = wb["TIMES"]
rows = list(ws.iter_rows(values_only=True))
hdr  = [str(c).strip() if c else "" for c in rows[0]]
lv_c = next(i for i,h in enumerate(hdr) if "레벨" in h or "level"   in h.lower())
tx_c = next(i for i,h in enumerate(hdr) if "본문" in h or "text"    in h.lower())
sc_c = next((i for i,h in enumerate(hdr) if "섹션" in h or "section" in h.lower()), None)
dt_c = next((i for i,h in enumerate(hdr) if "날짜" in h or "date"    in h.lower()), None)

pools = {"TIMES_L1": [], "TIMES_L2": [], "TIMES_L3": []}

for row in rows[1:]:
    if not row or len(row) <= max(lv_c, tx_c): continue
    lval = str(row[lv_c]).strip() if row[lv_c] else ""
    txt  = str(row[tx_c]).strip() if row[tx_c] else ""
    sec  = str(row[sc_c]).strip() if (sc_c and row[sc_c]) else ""
    if not lval or len(txt) < 50 or sec in EXCL: continue
    if len(re.findall(r"[A-Za-z']+", txt)) < 100: continue
    m = re.search(r"\d+", lval)
    if not m or m.group() not in ("1","2","3"): continue
    key = f"TIMES_L{m.group()}"
    if key not in LEVELS: continue
    cls = classify(txt, key)
    if cls.skip_cefr: continue
    a = avg_sl(txt)
    if a < 1.0: continue
    yr = None
    if dt_c is not None and len(row) > dt_c and row[dt_c]:
        ym = _YEAR_RE.search(str(row[dt_c]))
        if ym: yr = int(ym.group())
    if yr is not None and yr < 2024: continue
    pools[key].append(a)

# ── basic.xlsx TIMES L3 ───────────────────────────────────────────────────
basic_l3 = []
try:
    wb2  = openpyxl.load_workbook(BASIC, read_only=True, data_only=True)
    # 시트 목록 확인
    print("basic.xlsx 시트:", wb2.sheetnames)
    for sname in wb2.sheetnames:
        ws2  = wb2[sname]
        rows2 = list(ws2.iter_rows(values_only=True))
        if not rows2: continue
        hdr2  = [str(c).strip() if c else "" for c in rows2[0]]
        lv2   = next((i for i,h in enumerate(hdr2) if "레벨" in h or "level"   in h.lower()), None)
        tx2   = next((i for i,h in enumerate(hdr2) if "본문" in h or "text"    in h.lower()), None)
        me2   = next((i for i,h in enumerate(hdr2) if "매체" in h or "media"   in h.lower()), None)
        if lv2 is None or tx2 is None: continue
        for row in rows2[1:]:
            if not row or len(row) <= max(lv2, tx2): continue
            media = str(row[me2]).strip() if (me2 and row[me2]) else ""
            lval  = str(row[lv2]).strip() if row[lv2] else ""
            txt   = str(row[tx2]).strip() if row[tx2] else ""
            if not lval or len(txt) < 30: continue
            # TIMES L3 필터
            is_times = ("TIMES" in media.upper() or "times" in sname.lower() or "TIMES" in sname)
            m = re.search(r"\d+", lval)
            if not m: continue
            if is_times and m.group() == "3":
                a = avg_sl(txt)
                if a >= 1.0: basic_l3.append(a)
        if basic_l3: break  # 첫 번째 매칭 시트에서 수집
except Exception as e:
    print(f"basic.xlsx 오류: {e}")

# ── 출력 ────────────────────────────────────────────────────────────────────
print()
print("=" * 60)
print("articles.xlsx 현행(2024~) TIMES L1/L2/L3 avg 분포")
print("=" * 60)
for key in ["TIMES_L1", "TIMES_L2", "TIMES_L3"]:
    xs = pools[key]
    spec = LEVELS[key]
    n = len(xs)
    if not n: print(f"{key}: 데이터 없음"); continue
    p5  = pct(xs,  5); p10 = pct(xs, 10)
    p25 = pct(xs, 25); p50 = pct(xs, 50)
    p75 = pct(xs, 75); p90 = pct(xs, 90)
    print(f"\n{key}  n={n}  avg_min={spec.avg_min}")
    print(f"  p5={p5:.2f}  p10={p10:.2f}  p25={p25:.2f}  p50={p50:.2f}  p75={p75:.2f}  p90={p90:.2f}")

print()
print("=" * 60)
print(f"basic.xlsx TIMES_L3  n={len(basic_l3)}")
print("=" * 60)
if basic_l3:
    xs = basic_l3
    print(f"  p5={pct(xs,5):.2f}  p10={pct(xs,10):.2f}  p25={pct(xs,25):.2f}"
          f"  p50={pct(xs,50):.2f}  p75={pct(xs,75):.2f}  p90={pct(xs,90):.2f}")
    print(f"  min={min(xs):.2f}  max={max(xs):.2f}")
    print(f"  현재 avg_min=15.5 → basic p10={pct(xs,10):.2f} 기준으로 설정된 것")
else:
    print("  basic.xlsx에서 TIMES L3 데이터를 찾지 못했습니다.")
