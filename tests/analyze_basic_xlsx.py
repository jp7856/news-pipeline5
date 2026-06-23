"""basic.xlsx (5개 매체 시트)를 분석해 매체×레벨별 범위 통계를 산출한다.

사용: python tests/analyze_basic_xlsx.py <경로/basic.xlsx>
- 본문 끝 각주(*로 시작하는 줄, 한글 뜻풀이)는 단어 수 계산에서 제외
- 인라인 각주 마커(단어에 붙은 *)는 제거
- 비산문 포맷(Photo News, 대화·토론 등)은 수치 범위에서 빼고 포맷으로만 집계
- 백분위(p5/p25/median/p75/p95)까지 출력 → 겹치지 않는 range 설계 근거
"""

import re
import sys
import statistics
from collections import defaultdict

import openpyxl

NON_PROSE_HINTS = (
    "photo news", "did you know", "debate", "debating", "speak out",
    "talk talk", "think about it", "advice", "quiz", "puzzle",
)


def is_non_prose(section: str) -> bool:
    s = (section or "").strip().lower()
    return any(h in s for h in NON_PROSE_HINTS)


def clean_body(body: str) -> list[str]:
    """각주(*) 줄 제거, 단락 리스트 반환. 인라인 * 마커 제거."""
    if not body:
        return []
    paragraphs = []
    for block in re.split(r"\n\s*\n", str(body)):
        lines = [
            ln.strip() for ln in block.splitlines()
            if ln.strip() and not ln.strip().startswith("*")
        ]
        if lines:
            text = " ".join(lines)
            text = text.replace("*", "")  # 인라인 각주 마커
            paragraphs.append(text)
    return paragraphs


def word_count(text: str) -> int:
    # 한글이 섞인 토큰은 제외하고 영어 단어만 카운트
    tokens = text.split()
    return sum(1 for t in tokens if re.search(r"[A-Za-z]", t))


def sentence_lengths(text: str) -> list[int]:
    sentences = re.split(r"(?<=[.!?])\s+", text)
    return [word_count(s) for s in sentences if word_count(s) >= 2]


def pct(sorted_vals, p):
    if not sorted_vals:
        return 0
    k = (len(sorted_vals) - 1) * (p / 100)
    f = int(k)
    c = min(f + 1, len(sorted_vals) - 1)
    return sorted_vals[f] + (sorted_vals[c] - sorted_vals[f]) * (k - f)


def analyze_sheet(ws, media, out):
    # 헤더 인덱스
    hdr = {c.value: i for i, c in enumerate(ws[1], start=1)}
    col_level = hdr.get("레벨")
    col_section = hdr.get("섹션")
    col_body = hdr.get("본문")
    col_title = hdr.get("제목")

    stats = defaultdict(list)        # level -> [(words, avg_sl, n_par, n_sent)]
    sections = defaultdict(set)
    nonprose = defaultdict(set)

    for r in range(2, ws.max_row + 1):
        level_raw = ws.cell(r, col_level).value
        if not level_raw:
            continue
        level = str(level_raw).replace("LEVEL ", "L").replace("LEVEL", "L").strip()
        section = str(ws.cell(r, col_section).value or "").strip()
        body = ws.cell(r, col_body).value
        sections[level].add(section)
        if is_non_prose(section):
            nonprose[level].add(section)
            continue
        paragraphs = clean_body(body)
        text = " ".join(paragraphs)
        words = word_count(text)
        if words < 10:   # 빈/메뉴 행 방어
            continue
        slens = sentence_lengths(text)
        avg_sl = sum(slens) / len(slens) if slens else 0
        stats[level].append((words, avg_sl, len(paragraphs), len(slens)))

    out.append(f"\n{'='*70}\n{media}\n{'='*70}")
    for level in sorted(stats):
        rows = stats[level]
        ws_ = sorted(r[0] for r in rows)
        sl_ = sorted(r[1] for r in rows)
        ps_ = sorted(r[2] for r in rows)
        st_ = sorted(r[3] for r in rows)
        out.append(f"\n  [{level}]  기사 {len(rows)}건")
        out.append(
            f"    단어수    min {ws_[0]:>4} | p5 {pct(ws_,5):>5.0f} | p25 {pct(ws_,25):>5.0f} "
            f"| median {statistics.median(ws_):>5.0f} | p75 {pct(ws_,75):>5.0f} | p95 {pct(ws_,95):>5.0f} | max {ws_[-1]:>4}"
        )
        out.append(
            f"    문장길이  min {sl_[0]:>4.1f} | p5 {pct(sl_,5):>5.1f} | p25 {pct(sl_,25):>5.1f} "
            f"| median {statistics.median(sl_):>5.1f} | p75 {pct(sl_,75):>5.1f} | p95 {pct(sl_,95):>5.1f} | max {sl_[-1]:>4.1f}"
        )
        out.append(
            f"    단락수    min {ps_[0]:>4} | median {statistics.median(ps_):>5.0f} | max {ps_[-1]:>4}"
            f"    |  문장수 min {st_[0]} | median {statistics.median(st_):.0f} | max {st_[-1]}"
        )
        out.append(f"    산문 섹션: {sorted(sections[level] - nonprose[level])}")
        if nonprose[level]:
            out.append(f"    비산문 포맷: {sorted(nonprose[level])}")


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\jp\Desktop\basic.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for name in wb.sheetnames:
        analyze_sheet(wb[name], name, out)
    report = "\n".join(out)
    open(r"C:\Users\jp\work\_analysis_report.txt", "w", encoding="utf-8").write(report)
    print("report written")


if __name__ == "__main__":
    main()
