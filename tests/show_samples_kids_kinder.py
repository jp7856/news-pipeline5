"""샘플 출력 전용.
(1) KIDS L1/L2/L3 p5~p15 구간 기사 5개씩
(2) KINDER Think About It / Speak Out / My Diary 각 5개씩
"""
import sys, io, re, random
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\jp\work\news-pipeline5")

import openpyxl
from agents.sub_agents.article_classifier import classify

XLSX_PATH = r"C:\Users\jp\Desktop\기사\articles.xlsx"

EXCLUDE_SECTIONS_KIDS = {
    "Photo News", "NE You", "Cover",
    "Did You Know", "Did You Know?", "Debating", "Q & A", "Cartoon",
}

_SENT_END = re.compile(r'(?<=[.!?])\s+')

def avg_sent_len(text: str) -> float:
    parts = _SENT_END.split(text.strip())
    wcs = [len(re.findall(r"[A-Za-z']+", p)) for p in parts]
    wcs = [w for w in wcs if w >= 1]
    return sum(wcs) / len(wcs) if wcs else 0.0

def pct(d_sorted: list, p: float) -> float:
    if not d_sorted: return 0.0
    idx = (len(d_sorted) - 1) * p / 100
    lo  = int(idx)
    hi  = min(lo + 1, len(d_sorted) - 1)
    return d_sorted[lo] + (d_sorted[hi] - d_sorted[lo]) * (idx - lo)

random.seed(42)

# ══════════════════════════════════════════════════════════════════════════
# (1) KIDS L1/L2/L3  p5~p15 구간 샘플
# ══════════════════════════════════════════════════════════════════════════
print("=" * 72)
print("(1) KIDS L1 / L2 / L3  p5~p15 구간 ARTICLE 샘플 5개씩")
print("=" * 72)

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb["KIDS"]
rows = list(ws.iter_rows(values_only=True))
hdr  = [str(c).strip() if c else "" for c in rows[0]]
lv_col  = next(i for i,h in enumerate(hdr) if "레벨" in h or "level"   in h.lower())
tx_col  = next(i for i,h in enumerate(hdr) if "본문" in h or "text"    in h.lower())
sc_col  = next((i for i,h in enumerate(hdr) if "섹션" in h or "section" in h.lower()), None)
ttl_col = next((i for i,h in enumerate(hdr) if "제목" in h or "title"   in h.lower()), None)

# ARTICLE 풀 수집
kids_pool: dict[str, list] = {}
for row in rows[1:]:
    if not row: continue
    lv  = str(row[lv_col]).strip()  if row[lv_col]  else ""
    txt = str(row[tx_col]).strip()  if row[tx_col]  else ""
    sec = str(row[sc_col]).strip()  if (sc_col  and row[sc_col])  else ""
    ttl = str(row[ttl_col]).strip() if (ttl_col and row[ttl_col]) else ""
    if not lv or not txt: continue
    m = re.search(r"\d+", lv)
    if not m or int(m.group()) == 0: continue
    key = f"KIDS_L{m.group()}"
    if key not in ("KIDS_L1", "KIDS_L2", "KIDS_L3"): continue
    if len(re.findall(r"[A-Za-z']+", txt)) < 5: continue
    if sec in EXCLUDE_SECTIONS_KIDS: continue
    cls = classify(txt, key)
    if cls.skip_cefr: continue
    avg = avg_sent_len(txt)
    if avg < 1.0: continue
    kids_pool.setdefault(key, []).append((avg, ttl, cls.article_type.value, txt))

for key in ["KIDS_L1", "KIDS_L2", "KIDS_L3"]:
    items = kids_pool.get(key, [])
    avgs  = sorted(x[0] for x in items)
    n     = len(avgs)
    p5v   = pct(avgs, 5)
    p15v  = pct(avgs, 15)

    band = [(a, ttl, fmt, t) for a, ttl, fmt, t in items if p5v <= a <= p15v]
    if not band:
        band = [(a, ttl, fmt, t) for a, ttl, fmt, t in items if a <= p15v]
    band.sort(key=lambda x: x[0])

    step  = max(1, len(band) // 5)
    picks = [band[i] for i in range(0, min(len(band), step * 5), step)][:5]

    print(f"\n{'─'*72}")
    print(f"{key}  (전체 n={n}, p5={p5v:.1f}~p15={p15v:.1f} 구간 {len(band)}건 → 균등 {len(picks)}개)")
    print(f"{'─'*72}")
    for i, (avg, ttl, fmt, txt) in enumerate(picks, 1):
        preview = txt[:200].replace("\n", " ")
        print(f"\n[{i}] avg={avg:.1f}  포맷={fmt}  제목={ttl[:60] or '(없음)'}")
        print(f"     {preview}{'...' if len(txt)>200 else ''}")

# ══════════════════════════════════════════════════════════════════════════
# (2) KINDER 포맷 섹션 원문 샘플
# ══════════════════════════════════════════════════════════════════════════
print("\n\n" + "=" * 72)
print("(2) KINDER 포맷 섹션 원문 샘플  (본문 앞 300자)")
print("=" * 72)

TARGET_SECTIONS = ("Think About It", "Speak Out", "My Diary")

wb2 = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws2 = wb2["KINDER"]
rows2 = list(ws2.iter_rows(values_only=True))
hdr2  = [str(c).strip() if c else "" for c in rows2[0]]
tx2   = next(i for i,h in enumerate(hdr2) if "본문" in h or "text"  in h.lower())
sc2   = next((i for i,h in enumerate(hdr2) if "섹션" in h or "section" in h.lower()), None)
ttl2  = next((i for i,h in enumerate(hdr2) if "제목" in h or "title"  in h.lower()), None)

sec_pool: dict[str, list] = {s: [] for s in TARGET_SECTIONS}
for row in rows2[1:]:
    if not row: continue
    txt = str(row[tx2]).strip()  if row[tx2]  else ""
    sec = str(row[sc2]).strip()  if (sc2  and row[sc2])  else ""
    ttl = str(row[ttl2]).strip() if (ttl2 and row[ttl2]) else ""
    if sec not in TARGET_SECTIONS or not txt: continue
    sec_pool[sec].append((ttl, txt))

for sec in TARGET_SECTIONS:
    items  = sec_pool[sec]
    sample = random.sample(items, min(5, len(items)))
    print(f"\n{'─'*72}")
    print(f"KINDER — '{sec}'  (전체 {len(items)}건, {len(sample)}건 샘플)")
    print(f"{'─'*72}")
    for i, (ttl, txt) in enumerate(sample, 1):
        preview = txt[:300].replace("\n", " ")
        print(f"\n[{i}] 제목={ttl[:60] or '(없음)'}")
        print(f"     {preview}{'...' if len(txt)>300 else ''}")
