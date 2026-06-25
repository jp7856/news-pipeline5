"""JUNIOR M FK 분포 + 하단 avg 기사의 FK 확인 — avg_min vs fk_max 방향 결정용."""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\jp\work\news-pipeline5")

import textstat
import openpyxl
from agents.sub_agents.cefr_checker import validate
from agents.sub_agents.article_classifier import classify

XLSX_PATH = r"C:\Users\jp\Desktop\기사\articles.xlsx"

EXCLUDE_SECTIONS = {
    "Photo News", "Briefs", "Star Brief", "News in Brief",
    "Cartoon", "Did You Know", "Did You Know?", "Debating", "Cover", "Q & A",
    "NE You",
    "My Journal", "Book Review",       # TIMES 독자 기고·서평
    "Stories", "Story",                # TIMES 창작소설
    "Readings for Junior",             # TIMES 보충읽기
    "VoA Broadcast News",              # TIMES L3 방송 스크립트
}
MIN_WC = 100

def pct(xs, p):
    s = sorted(xs)
    return s[min(int(len(s) * p / 100), len(s) - 1)]

# ── 데이터 로드 ────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb["JUNIOR M"]
rows = list(ws.iter_rows(values_only=True))
hdr  = [str(c).strip() if c else "" for c in rows[0]]
lv_col = next(i for i,h in enumerate(hdr) if "레벨" in h or "level"   in h.lower())
tx_col = next(i for i,h in enumerate(hdr) if "본문" in h or "text"    in h.lower())
sc_col = next(i for i,h in enumerate(hdr) if "섹션" in h or "section" in h.lower())

# (avg, fk, sec, txt)
data: dict[str, list[tuple]] = {"JUNIORM_L1": [], "JUNIORM_L2": []}

for row in rows[1:]:
    if not row or len(row) <= max(lv_col, tx_col, sc_col): continue
    lv  = str(row[lv_col]).strip() if row[lv_col] else ""
    txt = str(row[tx_col]).strip() if row[tx_col] else ""
    sec = str(row[sc_col]).strip() if row[sc_col] else ""
    if not lv or len(txt) < 30: continue
    m = re.search(r"\d+", lv)
    if not m: continue
    key = f"JUNIORM_L{m.group()}"
    if key not in data: continue

    if sec in EXCLUDE_SECTIONS: continue
    wc = len(re.findall(r"[A-Za-z']+", txt))
    if wc < MIN_WC: continue
    if classify(txt, key).skip_cefr: continue

    r   = validate(txt, key)
    fk  = textstat.flesch_kincaid_grade(txt)
    data[key].append((r.avg_sentence_len, fk, sec, txt))

# ── 참고: JUNIOR 레벨 FK 분포 (cefr_checker 현행 fk_max) ─────────────────
# JUNIOR_L1 fk_max=9.0 / JUNIOR_L2 fk_max=10.5 / JUNIOR_L3 fk_max=11.5
# (spec 값, articles.xlsx 실측 아님 — 이 스크립트는 JUNIORM만 측정)

# ── 1. JUNIORM_L1/L2 FK 분포 ──────────────────────────────────────────────
print("=" * 72)
print("1. JUNIORM FK 분포 — 필터 후 ARTICLE 전체")
print("   (EXCLUDE_SECTIONS + wc<100 + BRIEF/DIALOGUE 제외)")
print("=" * 72)
print()

# JUNIOR 참고용 FK 분포도 articles.xlsx에서 뽑자
wb2 = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws_j = wb2["JUNIOR"]
rows_j = list(ws_j.iter_rows(values_only=True))
hdr_j  = [str(c).strip() if c else "" for c in rows_j[0]]
lv_j   = next(i for i,h in enumerate(hdr_j) if "레벨" in h or "level"   in h.lower())
tx_j   = next(i for i,h in enumerate(hdr_j) if "본문" in h or "text"    in h.lower())
sc_j   = next(i for i,h in enumerate(hdr_j) if "섹션" in h or "section" in h.lower())

junior_fk: dict[str, list[float]] = {}
for row in rows_j[1:]:
    if not row or len(row) <= max(lv_j, tx_j, sc_j): continue
    lv  = str(row[lv_j]).strip() if row[lv_j] else ""
    txt = str(row[tx_j]).strip() if row[tx_j] else ""
    sec = str(row[sc_j]).strip() if row[sc_j] else ""
    if not lv or len(txt) < 30: continue
    if sec in EXCLUDE_SECTIONS: continue
    wc = len(re.findall(r"[A-Za-z']+", txt))
    if wc < 80: continue
    m = re.search(r"\d+", lv)
    if not m: continue
    key = f"JUNIOR_L{m.group()}"
    if classify(txt, key).skip_cefr: continue
    junior_fk.setdefault(key, []).append(textstat.flesch_kincaid_grade(txt))

print(f"  {'레벨':<14} {'n':>5}  {'fk_max':>6}  {'p10':>6} {'p50':>6} {'p90':>6}  {'판정'}")
print(f"  {'─'*14} {'─'*5}  {'─'*6}  {'─'*6} {'─'*6} {'─'*6}  {'─'*20}")

# 현행 fk_max
FK_MAX = {
    "JUNIOR_L1": 9.0, "JUNIOR_L2": 10.5, "JUNIOR_L3": 11.5,
    "JUNIORM_L1": 10.0, "JUNIORM_L2": 10.0,
}

for key in ("JUNIOR_L1", "JUNIOR_L2", "JUNIOR_L3"):
    xs = junior_fk.get(key, [])
    if not xs: continue
    fkmax = FK_MAX[key]
    p10v = pct(xs, 10); p50v = pct(xs, 50); p90v = pct(xs, 90)
    reject = sum(1 for v in xs if v > fkmax)
    judge = f"p90({p90v:.1f}) {'>' if p90v > fkmax else '≤'} fk_max → 탈락 {reject}건({reject/len(xs)*100:.0f}%)"
    print(f"  {key:<14} {len(xs):>5}  {fkmax:>6.1f}  {p10v:>6.1f} {p50v:>6.1f} {p90v:>6.1f}  {judge}")

print()
for key in ("JUNIORM_L1", "JUNIORM_L2"):
    xs = [fk for _, fk, _, _ in data[key]]
    if not xs:
        print(f"  {key:<14} 데이터 없음")
        continue
    fkmax = FK_MAX[key]
    p10v = pct(xs, 10); p50v = pct(xs, 50); p90v = pct(xs, 90)
    reject = sum(1 for v in xs if v > fkmax)
    judge = f"p90({p90v:.1f}) {'>' if p90v > fkmax else '≤'} fk_max → 탈락 {reject}건({reject/len(xs)*100:.0f}%)"
    print(f"  {key:<14} {len(xs):>5}  {fkmax:>6.1f}  {p10v:>6.1f} {p50v:>6.1f} {p90v:>6.1f}  {judge}")

# ── 2. FK 탈락 기준 후보별 거부율 ────────────────────────────────────────
print()
print("=" * 72)
print("2. JUNIORM fk_max 후보별 탈락률 — 어느 값이 적절한지")
print("=" * 72)
for key in ("JUNIORM_L1", "JUNIORM_L2"):
    xs = [fk for _, fk, _, _ in data[key]]
    if not xs: continue
    print(f"\n  {key}  (n={len(xs)})")
    print(f"  {'fk_max':>6}  {'탈락건':>6}  {'탈락률':>7}")
    for cand in (8.5, 9.0, 9.5, 10.0, 10.5, 11.0, 11.5, 12.0):
        rej = sum(1 for v in xs if v > cand)
        print(f"  {cand:>6.1f}  {rej:>6}  {rej/len(xs)*100:>6.1f}%")

# ── 3. 하단 avg 기사 — avg/FK 같이 ───────────────────────────────────────
print()
print("=" * 72)
print("3. 하단 avg 기사 15건 — avg ↑ FK 확인 (문장 짧아도 어휘 어려운지)")
print("=" * 72)

for key in ("JUNIORM_L1", "JUNIORM_L2"):
    items = sorted(data[key], key=lambda x: x[0])   # avg 오름차순
    print(f"\n  ▶ {key} 최저 avg 15건")
    print(f"  {'#':>3}  {'avg':>5}  {'FK':>5}  {'섹션':<22}  본문 앞 100자")
    print(f"  {'─'*3}  {'─'*5}  {'─'*5}  {'─'*22}  {'─'*50}")
    for i, (avg, fk, sec, txt) in enumerate(items[:15], 1):
        preview = txt[:100].replace("\n", " ")
        print(f"  {i:>3}  {avg:>5.1f}  {fk:>5.1f}  {sec:<22}  {preview}")

# ── 4. FK 높은 상위 기사 10건 ────────────────────────────────────────────
print()
print("=" * 72)
print("4. FK 상위 기사 10건 — JUNIORM의 '어려운' 기사는 어떻게 생겼나")
print("=" * 72)

for key in ("JUNIORM_L1", "JUNIORM_L2"):
    items = sorted(data[key], key=lambda x: -x[1])  # FK 내림차순
    print(f"\n  ▶ {key} 최고 FK 10건")
    print(f"  {'#':>3}  {'avg':>5}  {'FK':>5}  {'섹션':<22}  본문 앞 100자")
    print(f"  {'─'*3}  {'─'*5}  {'─'*5}  {'─'*22}  {'─'*50}")
    for i, (avg, fk, sec, txt) in enumerate(items[:10], 1):
        preview = txt[:100].replace("\n", " ")
        print(f"  {i:>3}  {avg:>5.1f}  {fk:>5.1f}  {sec:<22}  {preview}")
