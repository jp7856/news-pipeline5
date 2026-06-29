"""JUNIOR_L3 분석: basic.xlsx 표본 + articles.xlsx 2024~ 분포 + 탈락 샘플."""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\jp\work\news-pipeline5")

import openpyxl
from agents.sub_agents.cefr_checker import LEVELS
from agents.sub_agents.article_classifier import classify

ARTICLES = r"C:\Users\jp\Desktop\기사\articles.xlsx"
BASIC    = r"C:\Users\jp\Desktop\기사\basic.xlsx"

EXCL = {
    "Photo News","Briefs","Star Brief","News in Brief","Cartoon",
    "Did You Know","Did You Know?","Debating","Cover","Q & A","NE You",
    "My Journal","Book Review","Stories","Story",
    "Readings for Junior","VoA Broadcast News","Think About It","My Diary",
    "Pros & Cons",
}
_SE      = re.compile(r"(?<=[.!?])\s+")
_YEAR_RE = re.compile(r"\b(20\d{2})\b")
YEAR_CUTOFF = 2024
AVG_MIN = LEVELS["JUNIOR_L3"].avg_min  # 11.5

def avg_sl(t):
    p = [len(re.findall(r"[A-Za-z']+", s)) for s in _SE.split(t.strip())]
    p = [w for w in p if w >= 1]
    return sum(p) / len(p) if p else 0.0

def pct(xs, p):
    s = sorted(xs)
    if not s: return 0.0
    idx = (len(s)-1)*p/100
    lo = int(idx); hi = min(lo+1, len(s)-1)
    return s[lo] + (s[hi]-s[lo])*(idx-lo)

# ── 1. basic.xlsx JUNIOR_L3 ───────────────────────────────────────────────
basic_l3 = []
try:
    wb2 = openpyxl.load_workbook(BASIC, read_only=True, data_only=True)
    print("basic.xlsx 시트:", wb2.sheetnames)
    for sname in wb2.sheetnames:
        ws2   = wb2[sname]
        rows2 = list(ws2.iter_rows(values_only=True))
        if not rows2: continue
        hdr2 = [str(c).strip() if c else "" for c in rows2[0]]
        lv2  = next((i for i,h in enumerate(hdr2) if "레벨" in h or "level" in h.lower()), None)
        tx2  = next((i for i,h in enumerate(hdr2) if "본문" in h or "text"  in h.lower()), None)
        me2  = next((i for i,h in enumerate(hdr2) if "매체" in h or "media" in h.lower()), None)
        if lv2 is None or tx2 is None: continue
        for row in rows2[1:]:
            if not row or len(row) <= max(lv2, tx2): continue
            media = str(row[me2]).strip() if (me2 and row[me2]) else ""
            lval  = str(row[lv2]).strip() if row[lv2] else ""
            txt   = str(row[tx2]).strip() if row[tx2] else ""
            if not lval or len(txt) < 30: continue
            is_junior = ("JUNIOR" in media.upper() or "junior" in sname.lower() or "JUNIOR" in sname)
            # JUNIOR M 제외
            is_junior_m = ("JUNIOR M" in media.upper() or "JUNIORM" in media.upper()
                           or "junior m" in sname.lower() or "juniorm" in sname.lower())
            if is_junior_m: continue
            m = re.search(r"\d+", lval)
            if not m: continue
            if is_junior and m.group() == "3":
                a = avg_sl(txt)
                if a >= 1.0: basic_l3.append(a)
        if basic_l3: break
except Exception as e:
    print(f"basic.xlsx 오류: {e}")

print()
print("=" * 60)
print(f"basic.xlsx JUNIOR_L3  n={len(basic_l3)}")
print("=" * 60)
if basic_l3:
    xs = basic_l3
    print(f"  p5={pct(xs,5):.2f}  p10={pct(xs,10):.2f}  p25={pct(xs,25):.2f}  p50={pct(xs,50):.2f}")
    print(f"  min={min(xs):.2f}  max={max(xs):.2f}")
    print(f"  현재 avg_min={AVG_MIN}")
else:
    print("  basic.xlsx에서 JUNIOR_L3 데이터를 찾지 못했습니다.")

# ── 2. articles.xlsx 2024~ JUNIOR_L3 분포 + 탈락 목록 ────────────────────
wb = openpyxl.load_workbook(ARTICLES, read_only=True, data_only=True)
ws = wb["JUNIOR"]
rows = list(ws.iter_rows(values_only=True))
hdr  = [str(c).strip() if c else "" for c in rows[0]]
lv_c = next(i for i,h in enumerate(hdr) if "레벨"  in h or "level"   in h.lower())
tx_c = next(i for i,h in enumerate(hdr) if "본문"  in h or "text"    in h.lower())
sc_c = next((i for i,h in enumerate(hdr) if "섹션"  in h or "section" in h.lower()), None)
dt_c = next((i for i,h in enumerate(hdr) if "날짜"  in h or "date"    in h.lower()), None)
ttl_c= next((i for i,h in enumerate(hdr) if "제목"  in h or "title"   in h.lower()), None)

pool  = []   # (avg, fmt, sec, ttl, txt) — 2024~ ARTICLE
fails = []   # avg < AVG_MIN

for row in rows[1:]:
    if not row or len(row) <= max(lv_c, tx_c): continue
    lval = str(row[lv_c]).strip() if row[lv_c] else ""
    txt  = str(row[tx_c]).strip() if row[tx_c] else ""
    sec  = str(row[sc_c]).strip() if (sc_c and row[sc_c]) else ""
    ttl  = str(row[ttl_c]).strip() if (ttl_c and row[ttl_c]) else ""
    if not lval or len(txt) < 50 or sec in EXCL: continue
    if len(re.findall(r"[A-Za-z']+", txt)) < 80: continue
    m = re.search(r"\d+", lval)
    if not m or m.group() != "3": continue
    cls = classify(txt, "JUNIOR_L3")
    if cls.skip_cefr: continue
    a = avg_sl(txt)
    if a < 1.0: continue
    yr = None
    if dt_c is not None and len(row) > dt_c and row[dt_c]:
        ym = _YEAR_RE.search(str(row[dt_c]))
        if ym: yr = int(ym.group())
    if yr is not None and yr < YEAR_CUTOFF: continue
    pool.append((a, cls.article_type.value, sec, ttl, txt))
    if a < AVG_MIN:
        fails.append((a, cls.article_type.value, sec, ttl, txt))

avgs = [x[0] for x in pool]
n = len(avgs)

print()
print("=" * 60)
print(f"articles.xlsx 2024~ JUNIOR_L3  n={n}  avg_min={AVG_MIN}")
print("=" * 60)
if avgs:
    print(f"  p5={pct(avgs,5):.2f}  p10={pct(avgs,10):.2f}  p25={pct(avgs,25):.2f}"
          f"  p50={pct(avgs,50):.2f}  p75={pct(avgs,75):.2f}")
    fail_n = len(fails)
    print(f"  탈락(avg<{AVG_MIN}): {fail_n}건 / {fail_n/n*100:.1f}%")

# ── 3. 탈락 샘플 ─────────────────────────────────────────────────────────
fails.sort(key=lambda x: x[0])
picks = fails[:12]
print()
print("=" * 60)
print(f"JUNIOR_L3 2024~ 탈락 전체 {len(fails)}건 / 출력 {len(picks)}개  (avg 오름차순)")
print("=" * 60)
for a, fmt, sec, ttl, txt in picks:
    preview = txt[:150].replace("\n", " ")
    print(f"[{a:.2f}] | {fmt} | {sec} | {preview}")
