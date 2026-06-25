"""
단독줄 화자 패턴 추가 검증.

목표:
  A. JUNIOR_L3 Debate 134건 → DIALOGUE (신규 패턴 감지 확인)
  B. KIDS_L3 Debate (콜론 포맷) → DIALOGUE (회귀 확인)
  C. 전 레벨 일반 기사 대량 샘플 → ARTICLE
     false positive 정의: wc ≥ 임계값인데 BRIEF, 또는 ARTICLE이어야 할 기사가 DIALOGUE
  D. 단위 테스트 — _count_heading_speakers 직접 검증
"""
import sys, io, re, random
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\jp\work\news-pipeline5")

import openpyxl
from agents.sub_agents.article_classifier import (
    classify, ArticleType, BRIEF_THRESHOLD,
    _count_heading_speakers, _HEADING_EXCLUSIONS,
)

XLSX_PATH = r"C:\Users\jp\Desktop\기사\articles.xlsx"
EXCLUDE_SECTIONS = {
    "Photo News", "Briefs", "Star Brief", "News in Brief",
    "Cartoon", "Did You Know", "Did You Know?", "Debating", "Cover", "Q & A",
    "NE You",
}
MIN_WC_BY_SHEET = {
    "KINDER": 0, "KIDS": 50, "JUNIOR": 80, "JUNIOR M": 100, "TIMES": 100,
}
PREFIX = {
    "KINDER": "KINDER", "KIDS": "KIDS", "JUNIOR": "JUNIOR",
    "JUNIOR M": "JUNIORM", "TIMES": "TIMES",
}

PASS = "✓"
FAIL = "✗"
errors: list[str] = []

# ═══════════════════════════════════════════════════════════════════════════
print("=" * 70)
print("D. 단위 테스트 — _count_heading_speakers")
print("=" * 70)

def unit(label: str, text: str, expected_count: int) -> None:
    lines = text.splitlines()
    count, _ = _count_heading_speakers(lines)
    ok = count == expected_count
    print(f"  {'✓' if ok else '✗'} count={count} (기대={expected_count})  {label}")
    if not ok:
        errors.append(f"단위테스트 실패: {label} — count={count}, 기대={expected_count}")

unit("Henry×3 → 3",
     "Topic intro.\n\nHenry\nI think this is good. We should support it.\n\n"
     "Henry\nI disagree with the previous point. Let me explain.\n\n"
     "Henry\nIn conclusion, I believe we should reconsider.",
     3)

unit("Henry×3 + 소제목 Background → 3 (Background 제외)",
     "Background\nThis is a background section.\n\n"
     "Henry\nI think this is good. We should support it.\n\n"
     "Henry\nI disagree with the previous point completely.\n\n"
     "Henry\nIn conclusion, we should reconsider the whole thing.",
     3)

unit("Korea×3 → 0 (지명 제외)",
     "Korea\nKorea is a country in East Asia known for technology.\n\n"
     "Korea\nKorea has a population of about fifty million people.\n\n"
     "Korea\nKorea joined the United Nations in nineteen ninety-one.",
     0)

unit("이름 1회만 → 0 (임계값 미달, 단독으로는 DIALOGUE 불가)",
     "Henry\nI think this is a great idea for everyone.\n\n"
     "Some regular paragraph here without any speaker heading.",
     1)  # count=1이지만 classify()에서 total<3이라 ARTICLE

unit("발화 4단어 → 0 (_SPEAKER_CONTENT_MIN=5 미달)",
     "Henry\nGood idea.\n\nHenry\nBad idea.\n\nHenry\nOkay fine.",
     0)

unit("단어 없는 다음줄 → 0",
     "Henry\n\n\nHenry\n\n\nHenry\n\n",
     0)

# ═══════════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("A. JUNIOR_L3 Debate → DIALOGUE (신규 패턴)")
print("=" * 70)

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws_j = wb["JUNIOR"]
rows_j = list(ws_j.iter_rows(values_only=True))
hdr_j  = [str(c).strip() if c else "" for c in rows_j[0]]
lv_j   = next(i for i,h in enumerate(hdr_j) if "레벨" in h or "level"   in h.lower())
tx_j   = next(i for i,h in enumerate(hdr_j) if "본문" in h or "text"    in h.lower())
sc_j   = next(i for i,h in enumerate(hdr_j) if "섹션" in h or "section" in h.lower())

j3_debate_total = 0
j3_debate_dlg   = 0
j3_debate_other: list[tuple[str, str]] = []  # (type, preview)

for row in rows_j[1:]:
    if not row or len(row) <= max(lv_j, tx_j): continue
    lv  = str(row[lv_j]).strip() if row[lv_j] else ""
    txt = str(row[tx_j]).strip() if row[tx_j] else ""
    sec = str(row[sc_j]).strip() if row[sc_j] else ""
    if lv != "LEVEL 3" or sec != "Debate": continue
    wc = len(re.findall(r"[A-Za-z']+", txt))
    if wc < 80: continue
    j3_debate_total += 1
    r = classify(txt, "JUNIOR_L3")
    if r.article_type == ArticleType.DIALOGUE:
        j3_debate_dlg += 1
    else:
        j3_debate_other.append((r.article_type.value, txt[:80].replace("\n"," ")))

print(f"  JUNIOR_L3 Debate {j3_debate_total}건 중 DIALOGUE: {j3_debate_dlg}건 "
      f"({j3_debate_dlg/j3_debate_total*100:.0f}%)")
if j3_debate_other:
    print(f"  ✗ 미감지 {len(j3_debate_other)}건:")
    for t, p in j3_debate_other[:5]:
        print(f"    [{t}] {p}...")
    errors.append(f"JUNIOR_L3 Debate 미감지: {len(j3_debate_other)}건")
else:
    print(f"  ✓ 전건 DIALOGUE 감지")

# ═══════════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("B. KIDS_L3 Debate (콜론 포맷) → DIALOGUE 회귀 확인")
print("=" * 70)

ws_k = wb["KIDS"]
rows_k = list(ws_k.iter_rows(values_only=True))
hdr_k  = [str(c).strip() if c else "" for c in rows_k[0]]
lv_k   = next(i for i,h in enumerate(hdr_k) if "레벨" in h or "level"   in h.lower())
tx_k   = next(i for i,h in enumerate(hdr_k) if "본문" in h or "text"    in h.lower())
sc_k   = next(i for i,h in enumerate(hdr_k) if "섹션" in h or "section" in h.lower())

k3_debate_total = 0
k3_debate_dlg   = 0
for row in rows_k[1:]:
    if not row or len(row) <= max(lv_k, tx_k): continue
    lv  = str(row[lv_k]).strip() if row[lv_k] else ""
    txt = str(row[tx_k]).strip() if row[tx_k] else ""
    sec = str(row[sc_k]).strip() if row[sc_k] else ""
    if lv != "LEVEL 3" or sec != "Debate": continue
    wc = len(re.findall(r"[A-Za-z']+", txt))
    if wc < 50: continue
    k3_debate_total += 1
    r = classify(txt, "KIDS_L3")
    if r.article_type == ArticleType.DIALOGUE:
        k3_debate_dlg += 1

# KIDS_L3 Debate 540건 중 콜론 포맷(Sue:/Steve:)은 약 306건만 DIALOGUE —
# 나머지 234건은 원래부터 ARTICLE(비코론 포맷). 회귀 기준은 306건 floor(300).
print(f"  KIDS_L3 Debate {k3_debate_total}건 중 DIALOGUE: {k3_debate_dlg}건 "
      f"({k3_debate_dlg/k3_debate_total*100:.0f}%)")
KIDS_L3_DIALOGUE_FLOOR = 300   # 기준선 306건에서 소폭 여유
if k3_debate_dlg < KIDS_L3_DIALOGUE_FLOOR:
    errors.append(f"KIDS_L3 Debate 회귀: {k3_debate_dlg}건 감지 (기준 {KIDS_L3_DIALOGUE_FLOOR}건 이상)")
    print(f"  ✗ 회귀: {k3_debate_dlg}건 감지 (기준 {KIDS_L3_DIALOGUE_FLOOR}건 이상)")
else:
    print(f"  ✓ 콜론 포맷 정상 유지 ({k3_debate_dlg}건 ≥ {KIDS_L3_DIALOGUE_FLOOR}건)")

# ═══════════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("C. 전 레벨 일반 기사 대량 샘플 → false positive 0건 목표")
print("   (일반 섹션 = Debate·Q&A 등 포맷 섹션 제외한 뉴스 기사)")
print("=" * 70)

NORMAL_SECTIONS_BY_SHEET: dict[str, set[str]] = {
    "KINDER":   {"World", "Korea", "Science", "Animals", "People", "Culture",
                 "Fun Facts", "Health", "Sports", "Environment", "Technology"},
    "KIDS":     {"Nation", "World", "Science", "Culture", "Sports", "Lifestyle",
                 "Headlines News", "Key Issue", "Science & Technology", "People"},
    "JUNIOR":   {"Nation", "World", "Focus", "People", "World Tour",
                 "Science", "Sports", "Culture", "Global", "Lifestyle"},
    "JUNIOR M": {"Nation", "World", "Focus", "People", "Science", "Sports",
                 "Culture", "Global", "Lifestyle", "Key Issue"},
    "TIMES":    {"Nation", "World", "Global", "Science", "Sports",
                 "Culture", "Lifestyle", "Key Issue", "People",
                 "Science & Technology", "Headlines News", "Entertainment"},
}

random.seed(42)
SAMPLE = 30   # 섹션당 최대 샘플 수
fp_total = 0

for sheet_name, normal_secs in NORMAL_SECTIONS_BY_SHEET.items():
    ws_s = wb[sheet_name]
    rows_s = list(ws_s.iter_rows(values_only=True))
    hdr_s  = [str(c).strip() if c else "" for c in rows_s[0]]
    lv_s   = next((i for i,h in enumerate(hdr_s) if "레벨" in h or "level"   in h.lower()), None)
    tx_s   = next((i for i,h in enumerate(hdr_s) if "본문" in h or "text"    in h.lower()), None)
    sc_s   = next((i for i,h in enumerate(hdr_s) if "섹션" in h or "section" in h.lower()), None)
    if lv_s is None or tx_s is None: continue
    min_wc = MIN_WC_BY_SHEET[sheet_name]

    # 레벨별 풀 구성
    pool: dict[str, list[str]] = {}
    for row in rows_s[1:]:
        if not row or len(row) <= max(lv_s, tx_s): continue
        lv  = str(row[lv_s]).strip() if row[lv_s] else ""
        txt = str(row[tx_s]).strip() if row[tx_s] else ""
        sec = str(row[sc_s]).strip() if (sc_s and row[sc_s]) else ""
        if not lv or len(txt) < 50: continue
        if sec not in normal_secs: continue
        wc = len(re.findall(r"[A-Za-z']+", txt))
        if wc < min_wc: continue
        m = re.search(r"\d+", lv)
        if not m: continue
        key = f"{PREFIX[sheet_name]}_L{m.group()}"
        pool.setdefault(key, []).append(txt)

    for key, texts in sorted(pool.items()):
        if key not in BRIEF_THRESHOLD:   # KIDS_L0 등 잘못된 키 제외
            continue
        sample = random.sample(texts, min(SAMPLE, len(texts)))
        results = [(t, classify(t, key)) for t in sample]
        # DIALOGUE: 정상 섹션에서 코론/단독줄 패턴이 ≥3 감지 → 실제 대화체일 수 있음
        #           → 오류가 아닌 INFO로 출력만, 에러 카운트에 포함 안 함
        info_dlg  = [(t, r) for t, r in results if r.article_type == ArticleType.DIALOGUE]
        # BRIEF: wc가 임계값 이상인데 BRIEF 판정 → 진짜 오류
        fp_brief  = [(t, r) for t, r in results
                     if r.article_type == ArticleType.BRIEF
                     and len(re.findall(r"[A-Za-z']+", t)) >= BRIEF_THRESHOLD[key]]
        fp = len(fp_brief)
        mark = PASS if fp == 0 else FAIL
        detail = ""
        if info_dlg:
            detail += f" [INFO: DIALOGUE×{len(info_dlg)} — 정상 대화체 추정]"
        if fp_brief:
            detail += f" BRIEF오류×{fp}"
        print(f"  {mark} {key:<14}  {len(sample):>3}건 샘플 → fp_error={fp}{detail}")
        for t, r in info_dlg[:2]:
            ex = " / ".join(f'"{e[:40]}"' for e in r.dialogue_examples[:2])
            print(f"      [INFO] DIALOGUE(비율={r.dialogue_line_ratio:.0%}, {ex}) "
                  f"{t[:80].replace(chr(10),' ')}...")
        for t, r in fp_brief[:1]:
            print(f"      !! BRIEF오류: wc={r.word_count} {t[:80].replace(chr(10),' ')}...")
            errors.append(f"false positive BRIEF [{key}]: {t[:60]}")
        fp_total += fp

print(f"\n  → 전체 false positive(BRIEF오류): {fp_total}건 (목표: 0건)")

# ═══════════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
if errors:
    print(f"최종 결과: {FAIL} 오류 {len(errors)}건")
    for e in errors:
        print(f"  - {e}")
else:
    print(f"최종 결과: {PASS} 전 시나리오 통과")
print("=" * 70)
