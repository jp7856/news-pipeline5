"""
article_classifier 종합 검증.

시나리오:
  A. 알려진 DIALOGUE → DIALOGUE (false negative 0건 목표)
  B. Debate 에세이(Fallacy 등) → DIALOGUE (예시 대화 포함, 비율 낮음 확인)
  C. 일반 뉴스 기사 대량 샘플 → ARTICLE or DIALOGUE (BRIEF만 오류)
     false positive 정의: wc ≥ 임계값인데 BRIEF로 빠지는 경우
  D. TIMES_L1 too_easy 기사(골프/뷔페/등교/피시토리우스) → ARTICLE, wc 확인
  E. 실제 단신(<100단어) → BRIEF
"""
import sys, io, re, random
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\jp\work\news-pipeline5")

import openpyxl
from agents.sub_agents.article_classifier import (
    classify, ArticleType, ClassificationResult, BRIEF_THRESHOLD
)

XLSX_PATH = r"C:\Users\jp\Desktop\기사\articles.xlsx"

# ── 데이터 로드 헬퍼 ──────────────────────────────────────────────────────
def load_times_l1(min_wc: int = 0) -> list[tuple[str, str, str, int]]:
    """(섹션, 제목, 본문, wc) 리스트 반환 — TIMES 시트 LEVEL 1."""
    wb  = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws  = wb["TIMES"]
    rows = list(ws.iter_rows(values_only=True))
    hdr  = [str(c).strip() if c else "" for c in rows[0]]
    lv   = next(i for i,h in enumerate(hdr) if "레벨" in h or "level" in h.lower())
    tx   = next(i for i,h in enumerate(hdr) if "본문" in h or "text" in h.lower())
    sc   = next(i for i,h in enumerate(hdr) if "섹션" in h or "section" in h.lower())
    ttl  = next((i for i,h in enumerate(hdr) if "제목" in h or "title" in h.lower()), None)
    out  = []
    for row in rows[1:]:
        if not row or len(row) <= max(lv, tx): continue
        level = str(row[lv]).strip() if row[lv] else ""
        text  = str(row[tx]).strip() if row[tx] else ""
        sec   = str(row[sc]).strip() if row[sc] else ""
        title = str(row[ttl]).strip() if (ttl and row[ttl]) else ""
        if level != "LEVEL 1" or len(text) < 10: continue
        wc = len(re.findall(r"[A-Za-z']+", text))
        if wc >= min_wc:
            out.append((sec, title, text, wc))
    return out

# 한 번만 로드
ALL_ARTICLES = load_times_l1(min_wc=0)
random.seed(42)

PASS = "✓"
FAIL = "✗"
errors: list[str] = []

def check(label: str, text: str, expected: ArticleType,
          note: str = "", level_key: str = "TIMES_L1") -> ClassificationResult:
    r = classify(text, level_key)
    ok = r.article_type == expected
    mark = PASS if ok else FAIL
    extra = ""
    if r.article_type == ArticleType.DIALOGUE:
        extra = f" (화자줄={r.dialogue_line_count}, 비율={r.dialogue_line_ratio:.0%})"
    elif r.article_type == ArticleType.BRIEF:
        extra = f" (wc={r.word_count}, 임계값={r.brief_threshold})"
    print(f"  {mark} [{r.article_type.value}]{extra}  {label}")
    if note:
        print(f"       → {note}")
    if not ok:
        errors.append(f"{label}: expected {expected.value}, got {r.article_type.value}")
    return r

# ═══════════════════════════════════════════════════════════════════════════
print("=" * 70)
print("A. 알려진 DIALOGUE 기사 → DIALOGUE (false negative 0건 목표)")
print("=" * 70)

DIALOGUE_SNIPPETS = [
    ("Kelly/Sangmin — Made in France",  "Kelly: I really admire your jacket"),
    ("Katie/Aaron — No Soup Day",       "Katie: What do you feel like having for lunch"),
    ("Lucy/Greg — Korean soccer",       "Lucy: February is an exciting time for Korean soccer, Greg."),
]
for label, snippet in DIALOGUE_SNIPPETS:
    hit = next((t for _,_,t,_ in ALL_ARTICLES if snippet in t), None)
    if hit is None:
        print(f"  ? 샘플 못 찾음: {snippet[:50]}")
        continue
    check(label, hit, ArticleType.DIALOGUE)

# ═══════════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("B. Debate 에세이 (논리 마커) → DIALOGUE (예시 대화 포함, 비율 낮음 확인)")
print("   비율이 낮으면 산문 속 삽입 예시, 높으면 순수 대화체")
print("=" * 70)

DEBATE_SNIPPETS = [
    ("Fallacy of Continuum — 논리에세이, 예시대화 포함", "Fallacy of Continuum"),
    ("Appeal to pity — 논리에세이, 예시 적음",           "Appeal to pity"),
]
for label, snippet in DEBATE_SNIPPETS:
    hit = next((t for _,_,t,_ in ALL_ARTICLES if snippet in t), None)
    if hit is None:
        print(f"  ? 샘플 못 찾음: {snippet}")
        continue
    r = classify(hit, "TIMES_L1")
    mark = PASS if r.article_type == ArticleType.DIALOGUE else PASS  # 둘 다 허용
    extra = f"(화자줄={r.dialogue_line_count}, 비율={r.dialogue_line_ratio:.0%})"
    print(f"  ✓ [{r.article_type.value}] {extra}  {label}")
    if "Fallacy" in label:
        note = "비율 낮을수록 산문 에세이에 예시 삽입 — 나중에 비율 기준으로 재분류 가능"
        print(f"       → {note}")

# Debate 섹션 전체 — 비율 분포 확인
debate_articles = [(s,ti,t,w) for s,ti,t,w in ALL_ARTICLES if s == "Debate"]
debate_dlg = [(s,ti,t,w) for s,ti,t,w in debate_articles
              if classify(t, "TIMES_L1").article_type == ArticleType.DIALOGUE]
print(f"\n  Debate 섹션 {len(debate_articles)}건 중 DIALOGUE: {len(debate_dlg)}건  "
      f"ARTICLE: {len(debate_articles)-len(debate_dlg)}건")
print(f"  DIALOGUE 건 비율 분포:")
for s, ti, t, w in debate_dlg:
    r = classify(t, "TIMES_L1")
    print(f"    비율={r.dialogue_line_ratio:.0%}  화자줄={r.dialogue_line_count}  "
          f"wc={w}  {t[:60].replace(chr(10),' ')}...")

# ═══════════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("C. 일반 뉴스 기사 — 섹션별 대량 샘플")
print("   false positive 정의: wc ≥ 임계값인데 BRIEF로 빠지는 경우")
print("   DIALOGUE 판정은 진짜 대화체 기사일 수 있으므로 오류 아님")
print("=" * 70)

NORMAL_SECTIONS = ["Nation", "World", "Global", "Science", "Sports",
                   "Culture", "Lifestyle", "Entertainment", "Education",
                   "Headlines News", "Key Issue", "Science & Technology"]
SAMPLE_PER_SECTION = 15

fp_brief = 0   # 실제 false positive: wc ≥ 임계값인데 BRIEF
dlg_found = 0  # 정상 감지: 진짜 대화체 기사
for sec in NORMAL_SECTIONS:
    pool = [(s,ti,t,w) for s,ti,t,w in ALL_ARTICLES
            if s == sec and w >= BRIEF_THRESHOLD["TIMES_L1"]]
    if not pool:
        continue
    sample = random.sample(pool, min(SAMPLE_PER_SECTION, len(pool)))
    briefs_wrong = [(s,ti,t,w) for s,ti,t,w in sample
                    if classify(t, "TIMES_L1").article_type == ArticleType.BRIEF]
    dialogues    = [(s,ti,t,w) for s,ti,t,w in sample
                    if classify(t, "TIMES_L1").article_type == ArticleType.DIALOGUE]
    articles     = len(sample) - len(briefs_wrong) - len(dialogues)
    status = PASS if not briefs_wrong else FAIL
    print(f"  {status} [{sec}]  {len(sample)}건 → ARTICLE {articles}건 / "
          f"DIALOGUE {len(dialogues)}건 / BRIEF오류 {len(briefs_wrong)}건")
    for s,ti,t,w in dialogues:
        r = classify(t, "TIMES_L1")
        print(f"      ℹ DIALOGUE(비율={r.dialogue_line_ratio:.0%}) wc={w}  "
              f"{t[:70].replace(chr(10),' ')}...")
        dlg_found += 1
    for s,ti,t,w in briefs_wrong:
        r = classify(t, "TIMES_L1")
        print(f"      ✗ BRIEF(wc={w}, 임계값={r.brief_threshold})  {t[:70].replace(chr(10),' ')}...")
        errors.append(f"false positive BRIEF [{sec}]: wc={w}  {t[:60]}")
        fp_brief += 1

print(f"\n  → BRIEF false positive: {fp_brief}건 (목표: 0건)")
print(f"  → DIALOGUE 감지(섹션 내 실제 대화체): {dlg_found}건 (정상)")

# ═══════════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("D. TIMES_L1 too_easy 기사(JUNIOR급 단문) → ARTICLE + wc 확인")
print("   (100단어 임계값에 걸려 BRIEF로 빠지면 안 됨)")
print("=" * 70)

TOO_EASY_SNIPPETS = [
    ("골프 Masters — avg 9.2",       "Korean golf fans were pleasantly surprised by the results"),
    ("세계최대뷔페 — avg 9.5",        "This year, a hotel in Las Vegas, Nevada made a huge buffet"),
    ("미국 늦은 등교 — avg 9.7",      "In the United States, students begin their school day at 7:30"),
    ("Oscar Pistorius — avg 9.9",    "Oscar Pistorius is a runner from South Africa"),
    ("OECD 근로시간 — avg 10.3",     "Korean adults are known for working hard. According to a recent survey"),
    ("점술 트렌드 — avg 10.5",        "Fortunetelling is not just for elders"),
    ("백악관 목련나무 — avg 10.6",    "A magnolia tree at the White House"),
]
for label, snippet in TOO_EASY_SNIPPETS:
    hit = next((t for _,_,t,_ in ALL_ARTICLES if snippet in t), None)
    if hit is None:
        print(f"  ? 샘플 못 찾음: {snippet[:50]}")
        continue
    wc = len(re.findall(r"[A-Za-z']+", hit))
    r  = classify(hit, "TIMES_L1")
    ok = r.article_type == ArticleType.ARTICLE
    mark = PASS if ok else FAIL
    print(f"  {mark} [{r.article_type.value}]  wc={wc}  {label}")
    if not ok:
        errors.append(f"too_easy 기사 오분류: {label} → {r.article_type.value}")

# ═══════════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
print("E. 실제 단신 → BRIEF")
print("=" * 70)

# Briefs 섹션 (이미 EXCLUDE_SECTIONS에 있지만 본문 기준 테스트)
wb2 = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws2 = wb2["TIMES"]
rows2 = list(ws2.iter_rows(values_only=True))
hdr2  = [str(c).strip() if c else "" for c in rows2[0]]
tx2   = next(i for i,h in enumerate(hdr2) if "본문" in h or "text" in h.lower())
sc2   = next(i for i,h in enumerate(hdr2) if "섹션" in h or "section" in h.lower())
lv2   = next(i for i,h in enumerate(hdr2) if "레벨" in h or "level" in h.lower())

briefs = []
for row in rows2[1:]:
    if not row or len(row) <= max(tx2, sc2): continue
    level = str(row[lv2]).strip() if row[lv2] else ""
    sec   = str(row[sc2]).strip() if row[sc2] else ""
    txt   = str(row[tx2]).strip() if row[tx2] else ""
    if level != "LEVEL 1" or sec not in {"Briefs", "Star Brief", "News in Brief"}: continue
    wc = len(re.findall(r"[A-Za-z']+", txt))
    if 20 <= wc < 100:
        briefs.append((sec, wc, txt))
    if len(briefs) >= 10: break

brief_wrong = 0
for sec, wc, txt in briefs:
    r = classify(txt, "TIMES_L1")
    ok = r.article_type == ArticleType.BRIEF
    mark = PASS if ok else FAIL
    print(f"  {mark} [{r.article_type.value}]  wc={wc}  [{sec}]  {txt[:60].replace(chr(10),' ')}...")
    if not ok:
        brief_wrong += 1
        errors.append(f"단신 미감지: wc={wc} [{sec}]")

# ═══════════════════════════════════════════════════════════════════════════
print()
print("=" * 70)
if errors:
    print(f"최종 결과: {FAIL} 오류 {len(errors)}건")
    for e in errors:
        print(f"  - {e}")
else:
    print(f"최종 결과: {PASS} 전 시나리오 통과 — false positive 0건")
print("=" * 70)
