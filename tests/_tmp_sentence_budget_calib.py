"""전 레벨 문장 예산 캘리브레이션 — basic.xlsx 정상 산문의
문단 수 / 문단당 문장 수 / 문장당 단어 분포 (L3 캘리브레이션과 동일 하이진).

하이진: EXCL 섹션(VoA 포함) 제외, year>=2024, KIDS LEVEL 0 제외,
article_classifier.skip_cefr(비산문) 제외.
"""
import io
import re
import statistics
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, r"C:\Users\jp\work\news-pipeline5")
import openpyxl
from agents.sub_agents.cefr_checker import LEVELS
from agents.sub_agents.article_classifier import classify

BASIC = r"C:\Users\jp\Desktop\기사\basic.xlsx"

EXCL = {
    "Photo News", "Briefs", "Star Brief", "News in Brief", "Cartoon",
    "Did You Know", "Did You Know?", "Debating", "Cover", "Q & A", "NE You",
    "My Journal", "Book Review", "Stories", "Story",
    "Readings for Junior", "VoA Broadcast News", "Think About It", "My Diary",
    "Pros & Cons",
}
_SE = re.compile(r"(?<=[.!?])\s+")
_YEAR = re.compile(r"\b(20\d{2})\b")

SHEET_KEY = {"KINDER": "KINDER", "KIDS": "KIDS", "JUNIOR": "JUNIOR",
             "TIMES": "TIMES", "JUNIOR M": "JUNIORM"}


def sent_words(text):
    """문장별 단어 수 리스트."""
    out = []
    for s in _SE.split(text.strip()):
        n = len(re.findall(r"[A-Za-z']+", s))
        if n >= 1:
            out.append(n)
    return out


def pct(xs, p):
    s = sorted(xs)
    if not s:
        return 0.0
    idx = (len(s) - 1) * p / 100
    lo = int(idx)
    hi = min(lo + 1, len(s) - 1)
    return s[lo] + (s[hi] - s[lo]) * (idx - lo)


wb = openpyxl.load_workbook(BASIC, read_only=True, data_only=True)
stats: dict[str, dict] = {}

for sname in wb.sheetnames:
    prefix = SHEET_KEY.get(sname)
    if not prefix:
        continue
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(c).strip() if c else "" for c in rows[0]]
    sc = hdr.index("섹션"); lv = hdr.index("레벨"); dt = hdr.index("날짜"); tx = hdr.index("본문")
    for row in rows[1:]:
        if not row or len(row) <= tx:
            continue
        sec = str(row[sc]).strip() if row[sc] else ""
        lval = str(row[lv]).strip() if row[lv] else ""
        txt = str(row[tx]).strip() if row[tx] else ""
        if not txt or sec in EXCL:
            continue
        m = re.search(r"\d+", lval)
        if not m:
            continue
        ln = m.group()
        if ln == "0":  # KIDS LEVEL 0 등 — 대상 아님
            continue
        ym = _YEAR.search(str(row[dt]) if row[dt] else "")
        if ym and int(ym.group()) < 2024:
            continue
        key = f"{prefix}_L{ln}"
        if key in LEVELS:
            cls = classify(txt, key)
            if cls.skip_cefr:  # 비산문(브리프/대화문 등) 제외
                continue
        sw = sent_words(txt)
        if len(sw) < 3:
            continue
        paras = [p for p in txt.split("\n\n") if p.strip()]
        spp = [len([w for w in sent_words(p)]) for p in paras]
        b = stats.setdefault(key, {"n": 0, "avg_sl": [], "sents": [], "paras": [],
                                   "spp": [], "std": []})
        b["n"] += 1
        b["avg_sl"].append(sum(sw) / len(sw))
        b["sents"].extend(sw)
        b["paras"].append(len(paras))
        b["spp"].extend(s for s in spp if s > 0)
        if len(sw) >= 2:
            b["std"].append(statistics.stdev(sw))

order = ["KINDER_L1", "KINDER_L2", "KIDS_L1", "KIDS_L2", "KIDS_L3",
         "JUNIOR_L1", "JUNIOR_L2", "JUNIOR_L3", "JUNIORM_L1", "JUNIORM_L2",
         "TIMES_L1", "TIMES_L2", "TIMES_L3"]
print(f"{'bucket':12} {'n':>4} | {'avg_sl p25/50/75':>20} | {'문장단어 p75/p90':>16} | "
      f"{'문단수 p50':>8} {'문장/문단 p50':>10} | {'sl표준편차 p25/50/75':>20}")
for key in order:
    b = stats.get(key)
    if not b:
        print(f"{key:12} 데이터 없음")
        continue
    a = b["avg_sl"]; s = b["sents"]; st = b["std"]
    print(f"{key:12} {b['n']:>4} | {pct(a,25):5.1f}/{pct(a,50):5.1f}/{pct(a,75):5.1f}      | "
          f"{pct(s,75):5.0f}/{pct(s,90):5.0f}      | {pct(b['paras'],50):>8.0f} "
          f"{pct(b['spp'],50):>10.1f} | {pct(st,25):5.2f}/{pct(st,50):5.2f}/{pct(st,75):5.2f}")
