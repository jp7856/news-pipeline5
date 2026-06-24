import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\jp\Desktop\기사\articles.xlsx", read_only=True, data_only=True)

for sname in wb.sheetnames:
    ws   = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    hdr  = [str(c).strip() if c else "" for c in rows[0]]
    sec_col = next((i for i,h in enumerate(hdr) if "섹션" in h or "section" in h.lower()), None)
    tx_col  = next((i for i,h in enumerate(hdr) if "본문" in h or "text" in h.lower()), None)
    if sec_col is None: continue

    sections = {}
    for row in rows[1:]:
        if not row or len(row) <= sec_col: continue
        sec = str(row[sec_col]).strip() if row[sec_col] else ""
        txt = str(row[tx_col]).strip() if (tx_col and row[tx_col]) else ""
        wc  = len(re.findall(r"[A-Za-z']+", txt))
        if sec:
            sections.setdefault(sec, []).append(wc)

    print(f"[{sname}]")
    for sec, wcs in sorted(sections.items()):
        avg = sum(wcs) // len(wcs)
        print(f"  {sec:<35} {len(wcs):>5}건  avg_wc={avg:>4}")
